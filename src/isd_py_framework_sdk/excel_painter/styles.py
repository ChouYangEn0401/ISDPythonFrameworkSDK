"""
excel_painter.styles
====================
Ergonomic wrappers over openpyxl's style objects plus the consolidated
"styled table" preset distilled from the project report builders.

Everything accepts plain ``"RRGGBB"`` hex strings (no ``#``) so callers never
import ``openpyxl.styles`` directly.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ── Primitive builders ──────────────────────────────────────────────────────

def fill(hex_color: str) -> PatternFill:
    """Solid background fill from a ``"RRGGBB"`` hex string."""
    return PatternFill("solid", fgColor=hex_color)


def font(
    *,
    color: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    size: Optional[float] = None,
    name: Optional[str] = None,
) -> Font:
    kw: dict = {"bold": bold, "italic": italic}
    if color is not None:
        kw["color"] = color
    if size is not None:
        kw["size"] = size
    if name is not None:
        kw["name"] = name
    return Font(**kw)


def side(color: str = "D9D9D9", style: str = "thin") -> Side:
    return Side(border_style=style, color=color)


def box(
    color: str = "D9D9D9",
    style: str = "thin",
    *,
    left: Optional[Side] = None,
    right: Optional[Side] = None,
    top: Optional[Side] = None,
    bottom: Optional[Side] = None,
) -> Border:
    """A four-sided border.  Per-side overrides win over ``color``/``style``."""
    base = side(color, style)
    return Border(
        left=left or base,
        right=right or base,
        top=top or base,
        bottom=bottom or base,
    )


def align(
    *,
    horizontal: Optional[str] = None,
    vertical: str = "top",
    wrap: bool = False,
) -> Alignment:
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


# ── Status colour maps ──────────────────────────────────────────────────────
# A status map is ``{cell_value: (background_hex, font_hex)}``.

StatusFills = dict[str, "tuple[str, str]"]

# Spreadsheet-classic green / amber / red / grey, matching the project reports.
STATUS_GREEN = ("C6EFCE", "006100")
STATUS_AMBER = ("FFEB9C", "9C6500")
STATUS_RED = ("FFC7CE", "9C0006")
STATUS_GREY = ("BFBFBF", "3F3F3F")


# ── Table preset ─────────────────────────────────────────────────────────────

@dataclass
class TableStyle:
    """Declarative style for :meth:`ExcelPainter.style_table`.

    Defaults reproduce the consolidated project look: dark-blue header band with
    bold white text, thin grey gridlines, frozen header row, and an autofilter.
    """

    header_fill: str = "305496"          # dark blue
    header_font_color: str = "FFFFFF"    # white
    header_bold: bool = True
    header_size: float = 11
    header_height: float = 22
    header_align: str = "center"         # horizontal

    border_color: Optional[str] = "D9D9D9"   # None → no borders
    body_valign: str = "top"

    freeze_header: bool = True
    autofilter: bool = True

    banded: bool = False                 # zebra-stripe body rows
    band_fill: str = "F2F2F2"

    # default body column width when a column is absent from the widths dict
    default_width: float = 18

    def header_font(self) -> Font:
        return font(color=self.header_font_color, bold=self.header_bold, size=self.header_size)

    def header_fill_obj(self) -> PatternFill:
        return fill(self.header_fill)

    def border_obj(self) -> Optional[Border]:
        return box(self.border_color) if self.border_color else None


# A couple of ready-made variants.
def blue_table() -> TableStyle:
    """The default dark-blue-header look."""
    return TableStyle()


def minimal_table() -> TableStyle:
    """No borders, light-grey header — for dense data dumps."""
    return TableStyle(header_fill="E7E6E6", header_font_color="000000", border_color=None)
