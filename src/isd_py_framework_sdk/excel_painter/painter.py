"""
excel_painter.painter
=====================
``ExcelPainter`` — a fluent, stays-open Excel styling engine.

The whole point: **load the workbook once, chain every operation, save once.**
The legacy helpers re-opened and re-saved the file on every single call; here a
workbook is held in memory and every method returns ``self`` so operations
compose::

    from isd_py_framework_sdk.excel_painter import ExcelPainter, TableStyle, STATUS_GREEN

    (
        ExcelPainter.from_dataframe(df, sheet_name="Results")
        .style_table(widths={"Name": 22, "說明": 80}, wrap_cols=["說明"], text_cols=["ISI_ID"])
        .fill_by_value("狀態", {"✅完成": STATUS_GREEN, "❌失敗": ("FFC7CE", "9C0006")})
        .auto_width(only_columns=[1, 2])
        .save("out.xlsx", safe=True)
    )

Column arguments accept either a 1-based integer **or** a header name (resolved
against ``header_row``).
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Callable, Iterable, Optional, Sequence, Union

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from . import highlight as _hl
from .styles import TableStyle, align, box, fill
from .width import auto_resize
from ._format_snapshot import SheetFormatSnapshot

ColumnRef = Union[int, str]


class ExcelPainter:
    """Fluent wrapper around an openpyxl workbook + its active sheet."""

    def __init__(self, workbook: Workbook, *, sheet: Optional[str] = None) -> None:
        self.wb = workbook
        self.ws = workbook[sheet] if sheet else workbook.active

    # ── Constructors ───────────────────────────────────────────────────────

    @classmethod
    def new(cls, sheet_name: str = "Sheet1") -> "ExcelPainter":
        """A fresh, empty workbook."""
        wb = Workbook()
        wb.active.title = sheet_name
        return cls(wb)

    @classmethod
    def open(cls, path: Union[str, Path], *, sheet: Optional[str] = None) -> "ExcelPainter":
        """Open an existing ``.xlsx`` file for in-place styling."""
        return cls(load_workbook(path), sheet=sheet)

    @classmethod
    def from_dataframe(
        cls,
        df,
        *,
        sheet_name: str = "Sheet1",
        index: bool = False,
    ) -> "ExcelPainter":
        """Start from a DataFrame written into a fresh sheet."""
        painter = cls.new(sheet_name)
        painter.write_dataframe(df, index=index)
        return painter

    # ── Sheet management ─────────────────────────────────────────────────────

    def use_sheet(self, name: str) -> "ExcelPainter":
        """Switch the active sheet to *name* (must exist)."""
        self.ws = self.wb[name]
        return self

    def add_sheet(self, name: str, *, activate: bool = True) -> "ExcelPainter":
        """Create a new sheet; make it active unless ``activate=False``."""
        ws = self.wb.create_sheet(title=name)
        if activate:
            self.ws = ws
        return self

    # ── Data ─────────────────────────────────────────────────────────────────

    def write_dataframe(
        self,
        df,
        *,
        start_row: int = 1,
        header: bool = True,
        index: bool = False,
    ) -> "ExcelPainter":
        """Write *df* into the active sheet starting at ``start_row``."""
        from openpyxl.utils.dataframe import dataframe_to_rows

        r = start_row
        for row in dataframe_to_rows(df, index=index, header=header):
            # openpyxl emits a stray ``[None]`` separator between the header and
            # data only when BOTH index and header are written — skip just that
            # one, never a legitimate single-column None data row.
            if index and header and row == [None]:
                continue
            for c, val in enumerate(row, 1):
                self.ws.cell(row=r, column=c, value=_to_native(val))
            r += 1
        return self

    def cell_value(self, row: int, col: ColumnRef, *, header_row: int = 1) -> Any:
        """Read a cell value (column may be a name or a 1-based index)."""
        return self.ws.cell(row=row, column=self._col(col, header_row)).value

    # ── Column resolution ─────────────────────────────────────────────────────

    def _col(self, col: ColumnRef, header_row: int = 1) -> int:
        """Resolve a column ref (header name or 1-based int) to an index."""
        if isinstance(col, int):
            return col
        for c in range(1, self.ws.max_column + 1):
            if str(self.ws.cell(row=header_row, column=c).value) == str(col):
                return c
        raise KeyError(f"Column header {col!r} not found in row {header_row}.")

    def _headers(self, header_row: int = 1) -> list[str]:
        return [
            self.ws.cell(row=header_row, column=c).value
            for c in range(1, self.ws.max_column + 1)
        ]

    # ── Region iteration ──────────────────────────────────────────────────────

    def for_each(
        self,
        axis: str,
        callback: Callable[..., None],
        *,
        fixed: int,
        rng: Optional[Iterable[int]] = None,
        **kwargs,
    ) -> "ExcelPainter":
        """Run ``callback(painter, row, col, **kwargs)`` along a row or column.

        ``axis="row"`` iterates rows down a fixed column; ``axis="col"`` iterates
        columns across a fixed row.  ``rng`` defaults to the full used extent.
        """
        if axis == "row":
            rng = rng or range(1, self.ws.max_row + 1)
            for i in rng:
                callback(self, i, fixed, **kwargs)
        elif axis == "col":
            rng = rng or range(1, self.ws.max_column + 1)
            for i in rng:
                callback(self, fixed, i, **kwargs)
        else:
            raise ValueError("axis must be 'row' or 'col'")
        return self

    # ── Fills ─────────────────────────────────────────────────────────────────

    def fill_cell(self, row: int, col: ColumnRef, color: str, *, header_row: int = 1) -> "ExcelPainter":
        self.ws.cell(row=row, column=self._col(col, header_row)).fill = fill(color)
        return self

    def fill_by_value(
        self,
        column: ColumnRef,
        mapping: dict,
        *,
        header_row: int = 1,
        default: Optional[Union[str, tuple]] = None,
    ) -> "ExcelPainter":
        """Fill each data cell in *column* by looking its value up in *mapping*.

        A mapping value may be a background hex string, or a ``(bg, font)`` tuple
        (the project status-colour convention — also colours the font).
        """
        from .styles import font as _font

        ci = self._col(column, header_row)

        def resolve(spec):
            if spec is None:
                return None, None
            if isinstance(spec, tuple):
                return spec[0], spec[1]
            return spec, None

        for r in range(header_row + 1, self.ws.max_row + 1):
            cell = self.ws.cell(row=r, column=ci)
            spec = mapping.get(str(cell.value), default)
            bg, fg = resolve(spec)
            if bg:
                cell.fill = fill(bg)
            if fg:
                cell.font = _font(color=fg, bold=True)
        return self

    def fill_where(
        self,
        column: ColumnRef,
        predicate: Callable[[Any], bool],
        color: str,
        *,
        header_row: int = 1,
    ) -> "ExcelPainter":
        """Fill data cells of *column* whose value satisfies *predicate*."""
        ci = self._col(column, header_row)
        for r in range(header_row + 1, self.ws.max_row + 1):
            cell = self.ws.cell(row=r, column=ci)
            if predicate(cell.value):
                cell.fill = fill(color)
        return self

    def color_scale(
        self,
        column: ColumnRef,
        *,
        low_color: str = "FFFFFF",
        high_color: str = "63BE7B",
        header_row: int = 1,
        min_value: Optional[float] = None,
        max_value: Optional[float] = None,
    ) -> "ExcelPainter":
        """Heat-map a numeric column: linearly blend ``low_color``→``high_color``.

        Non-numeric cells are skipped.  ``min_value``/``max_value`` override the
        auto-detected data range (useful for a fixed 0..100 scale).
        """
        ci = self._col(column, header_row)
        values: list[tuple[int, float]] = []
        for r in range(header_row + 1, self.ws.max_row + 1):
            v = self.ws.cell(row=r, column=ci).value
            try:
                values.append((r, float(v)))
            except (TypeError, ValueError):
                continue
        if not values:
            return self

        lo = min_value if min_value is not None else min(v for _, v in values)
        hi = max_value if max_value is not None else max(v for _, v in values)
        span = (hi - lo) or 1.0
        for r, v in values:
            t = max(0.0, min(1.0, (v - lo) / span))
            self.ws.cell(row=r, column=ci).fill = fill(_blend(low_color, high_color, t))
        return self

    # ── Widths / structure ─────────────────────────────────────────────────────

    def set_widths(self, widths: dict, *, header_row: int = 1) -> "ExcelPainter":
        """Set column widths from a ``{header_or_index: width}`` mapping."""
        for col, w in widths.items():
            self.ws.column_dimensions[get_column_letter(self._col(col, header_row))].width = w
        return self

    def auto_width(
        self,
        *,
        padding: int = 2,
        min_width: float = 4.0,
        max_width: Optional[float] = 80.0,
        only_columns: Optional[Iterable[int]] = None,
    ) -> "ExcelPainter":
        """CJK-aware auto-fit of column widths (see :func:`width.auto_resize`)."""
        auto_resize(
            self.ws,
            padding=padding,
            min_width=min_width,
            max_width=max_width,
            only_columns=only_columns,
        )
        return self

    def freeze(self, *, row: Optional[int] = None, col: Optional[int] = None) -> "ExcelPainter":
        """Freeze panes above ``row`` and/or left of ``col`` (1-based)."""
        fr = (row or 0) + 1
        fc = (col or 0) + 1
        self.ws.freeze_panes = self.ws.cell(row=fr, column=fc)
        return self

    def hide_columns(
        self,
        *,
        by_name: Optional[Sequence[str]] = None,
        by_index: Optional[Sequence[int]] = None,
        header_row: int = 1,
    ) -> "ExcelPainter":
        for name in by_name or []:
            letter = get_column_letter(self._col(name, header_row))
            self.ws.column_dimensions[letter].hidden = True
        for idx in by_index or []:
            self.ws.column_dimensions[get_column_letter(idx)].hidden = True
        return self

    def add_autofilter(self, *, header_row: int = 1) -> "ExcelPainter":
        last = get_column_letter(self.ws.max_column)
        self.ws.auto_filter.ref = f"A{header_row}:{last}{self.ws.max_row}"
        return self

    def number_format(
        self,
        columns: Union[ColumnRef, Sequence[ColumnRef]],
        fmt: str = "@",
        *,
        header_row: int = 1,
    ) -> "ExcelPainter":
        """Apply a number format (default ``"@"`` = text) to whole columns' data."""
        cols = [columns] if isinstance(columns, (int, str)) else list(columns)
        for col in cols:
            ci = self._col(col, header_row)
            for r in range(header_row + 1, self.ws.max_row + 1):
                self.ws.cell(row=r, column=ci).number_format = fmt
        return self

    def title_banner(
        self,
        text: str,
        *,
        row: int = 1,
        n_cols: Optional[int] = None,
        fill_color: str = "1F3864",
        font_color: str = "FFFFFF",
        size: float = 14,
        height: float = 30,
    ) -> "ExcelPainter":
        """Merge a bold title banner across the top *row* of the table."""
        from openpyxl.utils import get_column_letter as _gcl
        from .styles import font as _font

        n = n_cols or self.ws.max_column or 1
        self.ws.merge_cells(f"A{row}:{_gcl(n)}{row}")
        cell = self.ws.cell(row=row, column=1, value=text)
        cell.fill = fill(fill_color)
        cell.font = _font(color=font_color, bold=True, size=size)
        cell.alignment = align(horizontal="center", vertical="center")
        self.ws.row_dimensions[row].height = height
        return self

    def style_row(
        self,
        row: int,
        *,
        fill_color: Optional[str] = None,
        font_color: Optional[str] = None,
        bold: bool = False,
    ) -> "ExcelPainter":
        """Apply a fill / font to every used cell in *row* (e.g. a totals row)."""
        from .styles import font as _font

        for c in range(1, self.ws.max_column + 1):
            cell = self.ws.cell(row=row, column=c)
            if fill_color:
                cell.fill = fill(fill_color)
            if font_color or bold:
                cell.font = _font(color=font_color, bold=bold)
        return self

    # ── The big one: styled table ───────────────────────────────────────────────

    def style_table(
        self,
        style: Optional[TableStyle] = None,
        *,
        header_row: int = 1,
        widths: Optional[dict] = None,
        wrap_cols: Sequence[ColumnRef] = (),
        left_cols: Sequence[ColumnRef] = (),
        text_cols: Sequence[ColumnRef] = (),
    ) -> "ExcelPainter":
        """Apply the consolidated table look to the current sheet.

        Header band + borders + per-column alignment/wrap + frozen header +
        autofilter, all driven by a :class:`TableStyle`.  ``widths`` /
        ``wrap_cols`` / ``left_cols`` / ``text_cols`` accept header names or
        1-based indices.
        """
        style = style or TableStyle()
        ws = self.ws
        n_cols = ws.max_column
        n_rows = ws.max_row

        wrap_idx = {self._col(c, header_row) for c in wrap_cols}
        left_idx = {self._col(c, header_row) for c in left_cols}
        text_idx = {self._col(c, header_row) for c in text_cols}
        border = style.border_obj()

        # Header band
        hfill = style.header_fill_obj()
        hfont = style.header_font()
        halign = align(horizontal=style.header_align, vertical="center", wrap=True)
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = halign
            if border:
                cell.border = border
        ws.row_dimensions[header_row].height = style.header_height

        # Body
        for r in range(header_row + 1, n_rows + 1):
            banded = style.banded and (r - header_row) % 2 == 0
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = align(
                    horizontal="left" if c in left_idx else None,
                    vertical=style.body_valign,
                    wrap=c in wrap_idx,
                )
                if border:
                    cell.border = border
                if c in text_idx:
                    cell.number_format = "@"
                if banded:
                    cell.fill = fill(style.band_fill)

        if widths:
            self.set_widths(widths, header_row=header_row)
        if style.freeze_header:
            self.freeze(row=header_row)
        if style.autofilter:
            self.add_autofilter(header_row=header_row)
        return self

    # ── Rich-text highlighting ────────────────────────────────────────────────

    def highlight_lcs(
        self,
        target_col: ColumnRef,
        compare_col: ColumnRef,
        *,
        into_col: Optional[ColumnRef] = None,
        mode: str = "word",
        highlight_color: str = "00AA00",
        default_color: str = "000000",
        header_row: int = 1,
    ) -> "ExcelPainter":
        """Per row, highlight the LCS of ``target_col`` vs ``compare_col``.

        The coloured rich text is written back into ``into_col`` (defaults to
        ``target_col``).
        """
        tci = self._col(target_col, header_row)
        cci = self._col(compare_col, header_row)
        dci = self._col(into_col, header_row) if into_col is not None else tci
        for r in range(header_row + 1, self.ws.max_row + 1):
            target = str(self.ws.cell(row=r, column=tci).value or "")
            compare = str(self.ws.cell(row=r, column=cci).value or "")
            if not target or not compare:
                continue
            self.ws.cell(row=r, column=dci).value = _hl.lcs_rich_text(
                target, compare, mode=mode,
                highlight_color=highlight_color, default_color=default_color,
            )
        return self

    def highlight_common_words(
        self,
        target_col: ColumnRef,
        compare_col: ColumnRef,
        *,
        into_col: Optional[ColumnRef] = None,
        highlight_color: str = "00AA00",
        default_color: str = "000000",
        header_row: int = 1,
    ) -> "ExcelPainter":
        """Per row, highlight words of ``target_col`` shared with ``compare_col``."""
        tci = self._col(target_col, header_row)
        cci = self._col(compare_col, header_row)
        dci = self._col(into_col, header_row) if into_col is not None else tci
        for r in range(header_row + 1, self.ws.max_row + 1):
            target = str(self.ws.cell(row=r, column=tci).value or "")
            compare = str(self.ws.cell(row=r, column=cci).value or "")
            if not target or not compare:
                continue
            self.ws.cell(row=r, column=dci).value = _hl.common_words_rich_text(
                target, compare,
                highlight_color=highlight_color, default_color=default_color,
            )
        return self

    # ── Format snapshot ────────────────────────────────────────────────────────

    def snapshot(self) -> SheetFormatSnapshot:
        """Capture the active sheet's formatting for later :meth:`restore`."""
        return SheetFormatSnapshot.capture(self.ws)

    def restore(self, snap: SheetFormatSnapshot) -> "ExcelPainter":
        """Reapply a previously captured snapshot onto the active sheet."""
        snap.restore(self.ws)
        return self

    # ── Save ──────────────────────────────────────────────────────────────────

    def save(
        self,
        path: Union[str, Path],
        *,
        safe: bool = False,
        open_after: bool = False,
    ) -> Path:
        """Save the workbook.

        ``safe=True`` — if the target file is locked (open in Excel), save to a
        timestamped sibling instead of raising ``PermissionError``.
        ``open_after=True`` — open the saved file in the OS default app.
        """
        path = Path(path)
        if safe:
            path = _free_path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(path)
        if open_after:
            _open_in_os(path)
        return path


# ── Module helpers ───────────────────────────────────────────────────────────

def _blend(c1: str, c2: str, t: float) -> str:
    """Linear blend between two ``RRGGBB`` hex colours; ``t`` in [0, 1]."""
    a = (int(c1[0:2], 16), int(c1[2:4], 16), int(c1[4:6], 16))
    b = (int(c2[0:2], 16), int(c2[2:4], 16), int(c2[4:6], 16))
    r, g, bl = (round(a[i] + (b[i] - a[i]) * t) for i in range(3))
    return f"{r:02X}{g:02X}{bl:02X}"


def _free_path(path: Path) -> Path:
    """Return *path*, or a timestamped sibling if *path* is locked."""
    if not path.exists():
        return path
    try:
        with path.open("ab"):
            return path
    except PermissionError:
        from datetime import datetime
        ts = datetime.now().strftime("%H%M%S")
        return path.with_name(f"{path.stem}_{ts}{path.suffix}")


def _open_in_os(path: Path) -> None:
    import os
    import sys

    try:
        if sys.platform.startswith("win"):
            os.startfile(str(path))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}"')
    except Exception:  # noqa: BLE001 — opening is best-effort
        pass


def _to_native(val):
    """Convert numpy / pandas scalars to Python natives (NaN/NaT → None)."""
    if val is None:
        return None
    try:
        if isinstance(val, float) and val != val:  # NaN
            return None
        if hasattr(val, "item"):  # numpy scalar
            return val.item()
    except Exception:  # noqa: BLE001
        pass
    try:
        import pandas as pd

        if pd.isna(val):
            return None
    except (ImportError, TypeError, ValueError):
        pass
    return val
