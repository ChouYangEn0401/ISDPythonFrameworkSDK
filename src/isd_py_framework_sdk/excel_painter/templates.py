"""
excel_painter.templates
=======================
Ready-made professional report templates — one function call each.

The core API is fluent (:class:`ExcelPainter`); these functions wrap the most
common report shapes (distilled from real reporting projects) so a caller does
not have to assemble the chain by hand:

* :func:`status_report`             — colour-coded result table + summary sheet
* :func:`summary_statistics_report` — numeric stats with heat-map + totals row
* :func:`multi_sheet_report`        — several styled sheets + an overview sheet
* :func:`diff_highlight_report`     — per-row word-diff highlighting

Every template returns the saved :class:`~pathlib.Path` and is built entirely on
the public :class:`ExcelPainter` API.
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional, Sequence, Union

from .painter import ColumnRef, ExcelPainter
from .styles import (
    STATUS_AMBER,
    STATUS_GREEN,
    STATUS_GREY,
    STATUS_RED,
    StatusFills,
    TableStyle,
)


# ── Helpers ──────────────────────────────────────────────────────────────────

def auto_status_fills(values) -> StatusFills:
    """Guess a ``{status: (bg, font)}`` map from leading emoji / keywords.

    ✅ / 已 / 成功 / 完成 → green, ⚠ → amber, ❌ → red, otherwise grey.
    """
    out: StatusFills = {}
    for v in dict.fromkeys(str(x) for x in values):
        if v.startswith("✅") or any(k in v for k in ("成功", "完成", "已回填", "已解決")):
            out[v] = STATUS_GREEN
        elif v.startswith("⚠"):
            out[v] = STATUS_AMBER
        elif v.startswith("❌"):
            out[v] = STATUS_RED
        else:
            out[v] = STATUS_GREY
    return out


def _header_row_for(title: Optional[str]) -> int:
    return 2 if title else 1


# ── Templates ────────────────────────────────────────────────────────────────

def status_report(
    df,
    path: Union[str, Path],
    *,
    status_column: ColumnRef,
    status_fills: Optional[StatusFills] = None,
    title: Optional[str] = None,
    sheet_name: str = "結果",
    widths: Optional[dict] = None,
    wrap_cols: Sequence[ColumnRef] = (),
    text_cols: Sequence[ColumnRef] = (),
    left_cols: Sequence[ColumnRef] = (),
    summary: bool = True,
    style: Optional[TableStyle] = None,
    auto_width: bool = True,
    safe: bool = False,
    open_after: bool = False,
) -> Path:
    """A colour-coded result table with an optional status-count summary sheet.

    The classic operations report: styled table, the ``status_column`` cells
    coloured by ``status_fills`` (auto-guessed if omitted), and a second sheet
    tallying each status.
    """
    if status_fills is None:
        status_fills = auto_status_fills(df[status_column] if isinstance(status_column, str) else df.iloc[:, status_column - 1])

    hr = _header_row_for(title)
    painter = ExcelPainter.new(sheet_name)
    if title:
        painter.title_banner(title, row=1, n_cols=len(df.columns))
    painter.write_dataframe(df, start_row=hr)
    if auto_width:
        painter.auto_width()
    painter.style_table(
        style, header_row=hr,
        widths=widths, wrap_cols=wrap_cols, text_cols=text_cols, left_cols=left_cols,
    )
    painter.fill_by_value(status_column, status_fills, header_row=hr)

    if summary:
        counts = df[status_column].value_counts() if isinstance(status_column, str) \
            else df.iloc[:, status_column - 1].value_counts()
        import pandas as pd

        summ = pd.DataFrame({"狀態": counts.index.astype(str), "筆數": counts.values})
        painter.add_sheet("統計摘要")
        painter.write_dataframe(summ)
        painter.style_table(widths={"狀態": 30, "筆數": 12})
        painter.fill_by_value("狀態", status_fills)
        painter.use_sheet(sheet_name)

    return painter.save(path, safe=safe, open_after=open_after)


def summary_statistics_report(
    df,
    path: Union[str, Path],
    *,
    numeric_cols: Optional[Sequence[ColumnRef]] = None,
    label_col: Optional[ColumnRef] = None,
    title: Optional[str] = None,
    sheet_name: str = "統計",
    heatmap: bool = True,
    totals: bool = True,
    totals_label: str = "總計",
    low_color: str = "FFFFFF",
    high_color: str = "63BE7B",
    widths: Optional[dict] = None,
    style: Optional[TableStyle] = None,
    safe: bool = False,
    open_after: bool = False,
) -> Path:
    """A numeric statistics table: heat-mapped value columns + a totals row.

    ``numeric_cols`` defaults to every numeric column in *df*.  Each gets a
    white→green colour scale; a bold ``總計`` (sum) row is appended.
    """
    import pandas as pd

    if numeric_cols is None:
        numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]

    hr = _header_row_for(title)
    painter = ExcelPainter.new(sheet_name)
    if title:
        painter.title_banner(title, row=1, n_cols=len(df.columns))
    painter.write_dataframe(df, start_row=hr)
    painter.auto_width()

    if heatmap:
        for col in numeric_cols:
            painter.color_scale(col, low_color=low_color, high_color=high_color, header_row=hr)

    if totals:
        total_row = hr + len(df) + 1
        # label in the first column (or label_col)
        from openpyxl.utils import get_column_letter  # noqa: F401 (clarity)

        label_idx = painter._col(label_col, hr) if label_col is not None else 1
        painter.ws.cell(row=total_row, column=label_idx, value=totals_label)
        for col in numeric_cols:
            ci = painter._col(col, hr)
            painter.ws.cell(row=total_row, column=ci, value=float(df[col].sum()))

    painter.style_table(style, header_row=hr, widths=widths)
    if totals:
        painter.style_row(hr + len(df) + 1, fill_color="DDEBF7", bold=True)

    return painter.save(path, safe=safe, open_after=open_after)


def multi_sheet_report(
    sheets: "dict[str, object]",
    path: Union[str, Path],
    *,
    overview: bool = True,
    overview_title: str = "總覽",
    style: Optional[TableStyle] = None,
    per_sheet: Optional[dict] = None,
    safe: bool = False,
    open_after: bool = False,
) -> Path:
    """Write several DataFrames as styled sheets, plus an overview index sheet.

    ``sheets`` maps ``sheet_name -> DataFrame``.  ``per_sheet`` optionally maps
    ``sheet_name -> kwargs`` forwarded to ``style_table`` (widths, wrap_cols…).
    """
    import pandas as pd

    per_sheet = per_sheet or {}
    names = list(sheets)
    painter: Optional[ExcelPainter] = None

    for name in names:
        df = sheets[name]
        if painter is None:
            painter = ExcelPainter.new(name)
        else:
            painter.add_sheet(name)
        painter.write_dataframe(df)
        painter.auto_width()
        painter.style_table(style, **per_sheet.get(name, {}))

    if overview and painter is not None:
        ov = pd.DataFrame({
            "工作表": names,
            "列數": [len(sheets[n]) for n in names],
            "欄數": [sheets[n].shape[1] for n in names],
        })
        # Put overview first by creating it then moving it to index 0.
        painter.add_sheet(overview_title)
        painter.title_banner(overview_title, row=1, n_cols=3)
        painter.write_dataframe(ov, start_row=2)
        painter.style_table(header_row=2, widths={"工作表": 28, "列數": 10, "欄數": 10})
        painter.wb.move_sheet(overview_title, -(len(painter.wb.sheetnames) - 1))

    assert painter is not None, "sheets must be non-empty"
    return painter.save(path, safe=safe, open_after=open_after)


def diff_highlight_report(
    df,
    path: Union[str, Path],
    *,
    column_pairs: Sequence[tuple],
    method: str = "common_words",
    title: Optional[str] = None,
    sheet_name: str = "比對",
    highlight_color: str = "1F7A1F",
    default_color: str = "555555",
    widths: Optional[dict] = None,
    wrap_cols: Sequence[ColumnRef] = (),
    style: Optional[TableStyle] = None,
    safe: bool = False,
    open_after: bool = False,
) -> Path:
    """Highlight per-row text overlap between column pairs.

    ``column_pairs`` is a list of ``(target_col, compare_col)`` or
    ``(target_col, compare_col, into_col)``.  ``method`` is ``"common_words"``
    (set overlap) or ``"lcs"`` (order-aware).
    """
    hr = _header_row_for(title)
    painter = ExcelPainter.new(sheet_name)
    if title:
        painter.title_banner(title, row=1, n_cols=len(df.columns))
    painter.write_dataframe(df, start_row=hr)
    painter.auto_width()
    painter.style_table(style, header_row=hr, widths=widths, wrap_cols=wrap_cols)

    for pair in column_pairs:
        target, compare = pair[0], pair[1]
        into = pair[2] if len(pair) > 2 else None
        if method == "lcs":
            painter.highlight_lcs(
                target, compare, into_col=into, mode="word",
                highlight_color=highlight_color, default_color=default_color, header_row=hr,
            )
        else:
            painter.highlight_common_words(
                target, compare, into_col=into,
                highlight_color=highlight_color, default_color=default_color, header_row=hr,
            )

    return painter.save(path, safe=safe, open_after=open_after)
