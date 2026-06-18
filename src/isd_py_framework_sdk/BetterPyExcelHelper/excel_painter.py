from openpyxl import load_workbook
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.styles import PatternFill
from typing import List, Callable, Optional, Literal
from itertools import cycle
from src.BetterPyExcelHelper.df_data_processor import get_value_from_cell

## LCS (Longest Common Subsequence) Algorithm
def word_lcs(s1: str, s2: str) -> str:
    """
    計算兩個字串的最長公共子序列
    """
    m, n = len(s1), len(s2)
    dp = [["" for _ in range(n + 1)] for _ in range(m + 1)]
    for i in range(m):
        for j in range(n):
            if s1[i] == s2[j]:
                dp[i + 1][j + 1] = dp[i][j] + s1[i]
            else:
                dp[i + 1][j + 1] = max(dp[i][j + 1], dp[i + 1][j], key=len)
    return dp[m][n]
def sentence_lcs(s1: str, s2: str, case_sensitive=False) -> List[str]:
    import re
    def tokenize(text: str, case_sensitive=False) -> List[str]:
        if not case_sensitive:
            text = text.upper()
        return re.findall(r'\b[\w&]+\b', text)

    a_tokens = tokenize(s1, case_sensitive)
    b_tokens = tokenize(s2, case_sensitive)

    n, m = len(a_tokens), len(b_tokens)
    dp = [[""] * (m + 1) for _ in range(n + 1)]

    for i in range(n):
        for j in range(m):
            if a_tokens[i] == b_tokens[j]:
                dp[i + 1][j + 1] = dp[i][j] + " " + a_tokens[i]
            else:
                dp[i + 1][j + 1] = max(dp[i][j + 1], dp[i + 1][j], key=len)
    return dp[n][m].strip().split()

## Helper for Coloring Cells
def _color_text_by_lcs(sheet, row: int, col: int, main_str: str, lcs_result: str,
                       highlight_color: str, default_color: str, mode: Literal["letter", "sentence"] = "letter"):
    """
    根據 LCS 結果，將主要字串中的字母或詞語標記顏色
    """
    text_blocks: List[TextBlock] = []

    if mode == "letter":
        lcs_chars = list(lcs_result)
        for ch in main_str:
            if lcs_chars and ch == lcs_chars[0]:
                text_blocks.append(TextBlock(InlineFont(color=highlight_color, b=True), ch))
                lcs_chars.pop(0)
            else:
                text_blocks.append(TextBlock(InlineFont(color=default_color, i=True), ch))

    elif mode == "sentence":
        words = main_str.split(' ')
        lcs_words = set(lcs_result.split())
        for word in words:
            color = highlight_color if word in lcs_words else default_color
            text_blocks.append(TextBlock(InlineFont(color=color), word + " "))  # 保留句子間隔

    else:
        raise ValueError(f"Unknown mode: {mode}")

    # 設定儲存格的內容為帶有顏色的文本
    cell = sheet.cell(row=row, column=col)
    cell.value = CellRichText(text_blocks)

def _color_text_by_jaccard_set(sheet, row: int, col: int,
                               main_str: str, highlight_words: list[str],
                               highlight_color: str = "00AA00",
                               default_color: str = "000000"):
    """
    將 main_str 的內容依據 highlight_words 設定不同顏色，顯示在 (row, col)
    """
    text_blocks: List[TextBlock] = []
    highlight_set = set(w.lower() for w in highlight_words)

    words = main_str.split()
    for word in words:
        color = highlight_color if word.lower() in highlight_set else default_color
        text_blocks.append(TextBlock(InlineFont(color=color), word + " "))  # 加空格保留格式

    cell = sheet.cell(row=row, column=col)
    cell.value = CellRichText(text_blocks)



## BG Fill
def set_cell_fill_color(sheet, row: int, col: int, hex_color: str = "FFFF00"):
    """
    設定指定儲存格的底色(背景顏色)
    :param sheet: 工作表
    :param row: 第幾列（從 1 開始）
    :param col: 第幾欄（從 1 開始）
    :param hex_color: 六碼十六進位色碼（不含 #）
    """
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    sheet.cell(row=row, column=col).fill = fill


## Region Manipulator
def entire_axis_run(
        sheet,
        paint_iteration_axis: Literal["row", "col"] = "row",         # "row" 或 "col"
        paint_range: Optional[range] = None,         # 要跑哪些 row 或 col
        callback: Callable[..., None] = None,        # callback(sheet, row, col, **kwargs)
        fixed_col: Optional[int] = None,             # 若 mode 是 row，指定固定 col
        fixed_row: Optional[int] = None,             # 若 mode 是 col，指定固定 row
        **kwargs
    ):
    """
    通用處理器，針對 row 或 col 執行 callback(sheet, row, col, **kwargs)

    :param sheet: openpyxl 工作表
    :param paint_iteration_axis: 'row' 或 'col'
    :param paint_range: 要處理的 row/col 範圍（如 range(2, 10)）
    :param callback: 要執行的 callback(sheet, row, col, **kwargs)
    :param fixed_col: mode=row 時要處理的欄位
    :param fixed_row: mode=col 時要處理的列
    :param kwargs: 傳給 callback 的額外參數
    """
    if paint_iteration_axis not in ("row", "col"):
        raise ValueError("mode 必須是 'row' 或 'col'")

    if paint_iteration_axis == "row" and fixed_col is None:
        raise ValueError("mode 為 'row' 時必須指定 fixed_col")
    if paint_iteration_axis == "col" and fixed_row is None:
        raise ValueError("mode 為 'col' 時必須指定 fixed_row")

    max_index = sheet.max_row if paint_iteration_axis == "row" else sheet.max_column
    paint_range = paint_range or range(1, max_index + 1)

    for idx in paint_range:
        row = idx if paint_iteration_axis == "row" else fixed_row
        col = fixed_col if paint_iteration_axis == "row" else idx
        callback(sheet, row, col, **kwargs)

# untested
def highlight_column_header(sheet, row, color="FFFF00"):
    """ 全欄上色，基於顏色輪替處理邏輯 """
    column_colors = cycle([color])  # 只用這個顏色
    entire_axis_run(sheet, paint_iteration_axis="col", custom_paint_range=range(2, sheet.max_row + 1),
                    callback=set_cell_fill_color, row=row, hex_color=next(column_colors))
# untested
def row_runner(sheet, col, colors=None):
    """ 全列上色，顏色選擇邏輯（輪替或單色） """
    if colors is None:
        colors = ["FFFF00"]  # 預設顏色

    column_colors = cycle(colors)  # 如果有多個顏色，會輪替
    for r in range(1, sheet.max_column + 1):
        color = next(column_colors)  # 取得當前顏色
        set_cell_fill_color(sheet, row=r, col=col, hex_color=color)

## Column Color Painter Utilities
def renamerename__make_lcs_col_painter(
            main_str_col: int,
            cmp_str_col: int,
            *,
            highlight_color: str = "00AA00",
            default_color: str = "000000"
    ) -> Callable:
    """
    根據 LCS 結果，建立一個上色的 callback 函數

    回傳一個 callback，可用於 entire_axis_run，從指定欄位讀資料後上色。
    :param main_str_col: 主要字串所在的欄位（main column）
    :param cmp_str_col: LCS 結果所在的欄位（LCS column）
    :param highlight_color: 上色的顏色，當字元在 LCS 中時使用
    :param default_color: 預設顏色，當字元不在 LCS 中時使用
    :return: 回傳一個 callback 函數，這個函數將被 `entire_axis_run` 呼叫來處理儲存格顏色
    """
    def _callback(sheet, row, col):  # col 不用管
        try:
            target_str = get_value_from_cell(sheet, row=row, col=col, expected_type=str, none_value_replacement="")
            main_str = get_value_from_cell(sheet, row=row, col=main_str_col, expected_type=str, none_value_replacement="")
            cmp_str = get_value_from_cell(sheet, row=row, col=cmp_str_col, expected_type=str, none_value_replacement="")
            lcs_str = ' '.join(sentence_lcs(main_str, cmp_str))
            if not lcs_str:
                return
            _color_text_by_lcs(sheet, row, col=col,
                               main_str=target_str,
                               lcs_result=lcs_str,
                               highlight_color=highlight_color,
                               default_color=default_color,
                               mode="sentence")
        except ValueError as e:
            print(f"Error processing sheet={sheet.title}, row={row}, col={cmp_str_col}: {e}")

    return _callback

def renamerename__make_jaccard_col_painter(
        main_str_col: int,
        compare_col: int,
        *,
        highlight_color: str = "00AA00",
        default_color: str = "000000"
    ) -> Callable:
    def _callback(sheet, row, col):
        try:
            target_str = get_value_from_cell(sheet, row=row, col=col, expected_type=str, none_value_replacement="")
            main_str = get_value_from_cell(sheet, row=row, col=main_str_col, expected_type=str, none_value_replacement="")
            compare_str = get_value_from_cell(sheet, row=row, col=compare_col, expected_type=str, none_value_replacement="")
            if not compare_str:
                return

            common_words = get_common_words_jaccard(main_str, compare_str)
            _color_text_by_jaccard_set(sheet, row=row, col=col,  # ✅ 上色目標是 compare_col
                                       main_str=target_str,
                                       highlight_words=common_words,
                                       highlight_color=highlight_color,
                                       default_color=default_color)
        except (ValueError, TypeError) as e:
            print(f"Error processing sheet={sheet.title}, row={row}, col={col}: {e}")

    return _callback

def make_fill_conditional_callback(condition_colors: dict, default_colors: List[str] = None) -> Callable:
    """
    根據欄位的文字條件設定填色
    :param condition_colors: {'收錄': 'CCFFCC', '不收錄': 'FFCCCC'}
    :param default_colors: 其他情況的顏色，例如 ['EEEEEE']
    :return: callback function 用於 entire_axis_run
    """
    def _callback(sheet, row, col, **kwargs):
        val = sheet.cell(row=row, column=col).value
        if val in condition_colors:
            color = condition_colors[val]
        else:
            color = default_colors[0] if default_colors else "FFFFFF"  # fallback 白色
        set_cell_fill_color(sheet, row, col, color)
    return _callback

def get_common_words_jaccard(str1: str, str2: str, case_sensitive: bool = False) -> set:
    """
    計算 Jaccard 相似度，並返回相同的詞彙
    """
    if not case_sensitive:
        str1 = str1.lower()
        str2 = str2.lower()
    words1 = set(str1.split())
    words2 = set(str2.split())
    return words1 & words2


## Basic Load File
def load_excel(filename, sheetname = ""):
    wb = load_workbook(filename)
    return wb, (wb[sheetname] if sheetname!="" else wb.active)
def save_excel(filename, wb):
    wb.save(filename)

## Region Manipulation Functions
def entire_axis_run(sheet, paint_iteration_axis: Literal["row", "col"], custom_paint_range: Optional[range],
                    callback: Callable, fixed_col: Optional[int] = None, fixed_row: Optional[int] = None, **kwargs):
    """
    在整個行或列中執行指定的 callback 函數
    """
    if paint_iteration_axis not in ("row", "col"):
        raise ValueError("Iteration axis must be 'row' or 'col'")
    if paint_iteration_axis == "row" and fixed_col is None:
        raise ValueError("IterationAxis=row, `fix-col` must be a value")
    elif paint_iteration_axis == "col" and fixed_row is None:
        raise ValueError("IterationAxis=col, `fix-row` must be a value")

    max_index = sheet.max_row if paint_iteration_axis == "row" else sheet.max_column
    paint_range = custom_paint_range or range(1, max_index + 1)

    for idx in paint_range:
        row = fixed_row if paint_iteration_axis == "col" else idx
        col = fixed_col if paint_iteration_axis == "row" else idx
        callback(sheet, row, col, **kwargs)


## Example Usage
def highlight_column_headers(sheet, row, color="FFFF00"):
    """
    高亮顯示某列標題
    """
    entire_axis_run(sheet, paint_iteration_axis="col", custom_paint_range=range(2, sheet.max_row + 1), callback=set_cell_fill_color, row=row, hex_color=color)

