import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

def open_file_when_program_end(output_file) -> bool:
    """開啟 Excel 檔案（Windows）。"""
    try:
        os.system(f"start \"\" \"{output_file}\"")  # 在路徑包含空格時，需要用雙引號包起來
        print(f"⏩ 已嘗試開啟檔案 '{output_file}'。")
        return True
    except Exception as e:
        print(f"❌ 開啟檔案時發生錯誤：{e}")
        return False

def get_display_width(text: str) -> int:
    import wcwidth
    if text is None:
        return 0
    return sum(wcwidth.wcwidth(c) for c in str(text))

def resize_excel_columns(
        output_file_path: str,
        sheet_name: str = "Sheet1",
        include_index: bool = False
    ) -> bool:
    """調整 Excel 欄位寬度。"""
    try:
        workbook = load_workbook(output_file_path)
        sheet = workbook[sheet_name]

        df = pd.read_excel(output_file_path, sheet_name=sheet_name, index_col=0 if include_index else None)
        header = [df.index.name] + list(df.columns) if include_index else list(df.columns)

        for i, column_cells in enumerate(sheet.columns):
            # 計算欄位名稱（header）寬度
            header_width = get_display_width(header[i]) if i < len(header) else 0

            # 計算資料欄最大寬度
            content_widths = [
                get_display_width(cell.value) for cell in column_cells if cell.value is not None
            ]
            max_content_width = max(content_widths, default=0)

            # 取 header 與內容最大值，加上額外 padding
            adjusted_width = max(header_width + 5, max_content_width)

            # 設定 Excel 欄位寬度
            col_letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[col_letter].width = adjusted_width

        workbook.save(output_file_path)
        print(f"⏩ 檔案 '{output_file_path}', 工作表 '{sheet_name}' 的欄位寬度已自動調整。")
        return True
    except FileNotFoundError:
        print(f"❌ 錯誤：找不到檔案 '{output_file_path}'，無法進行欄寬調整。")
        return False
    except KeyError:
        print(f"❌ 錯誤：工作表 '{sheet_name}' 不存在於檔案 '{output_file_path}' 中，無法進行欄寬調整。")
        return False
    except Exception as e:
        print(f"❌ 調整欄寬時發生錯誤：{e}")
        return False
    except Exception as e:
        print(f"❌ 欄寬調整失敗：{e}")
        return False

def save_dataframe_to_excel(
        df: pd.DataFrame,
        output_file_path: str,
        sheet_name: str = "Sheet1",
        include_index: bool = False
    ) -> bool:
    """儲存 DataFrame 到 Excel，不包含額外功能。"""
    try:
        df.to_excel(output_file_path, sheet_name=sheet_name, index=include_index)
        print(f"⏩ DataFrame 已儲存到 '{output_file_path}', 工作表 '{sheet_name}'。")
        return True
    except Exception as e:
        print(f"❌ 儲存 Excel 時發生錯誤：{e}")
        return False


def save_dataframe_entire_workflow(
        df: pd.DataFrame,
        output_file: str,
        sheet_name: str = "Sheet1",
        auto_resize: bool = True,
        open_file: bool = False,
        include_index: bool = False,
        hidden_columns_by_name: list[str] = None, hidden_columns_by_index: list[int] = None,
        freeze_the_row: int = None, freeze_the_col: int = None,
    ):
    """
    將 DataFrame 儲存到 Excel 檔案，並可選擇自動調整欄寬、開啟檔案和包含索引。

    Args:
        df (pd.DataFrame): 要儲存的 DataFrame。
        output_file (str): 輸出的 Excel 檔案路徑。
        sheet_name (str, optional): 工作表名稱。預設為 "Sheet1"。
        auto_resize (bool, optional): 是否自動調整欄寬。預設為 True。
        open_file (bool, optional): 是否在儲存後開啟檔案。預設為 False。
        include_index (bool, optional): 是否包含 DataFrame 的索引寫入 Excel。預設為 False。
    """
    try:
        flag = save_dataframe_to_excel(df, output_file, sheet_name, include_index)
        if not flag:
            print("⚠️ Program Stop At `save_dataframe_to_excel`")
            return False

        if auto_resize:
            flag = resize_excel_columns(output_file, sheet_name, include_index)
            if not flag:
                print("⚠️ Program Stop At `resize_excel_columns`")
                return False

        if hidden_columns_by_name or hidden_columns_by_index:
            flag = hide_excel_columns(output_file, sheet_name, hidden_columns_by_name, hidden_columns_by_index)
            if not flag:
                print("⚠️ Program Stop At `hide_excel_columns`")
                return False

        if freeze_the_row or freeze_the_col:
            flag = freeze_excel_panes(output_file, sheet_name, freeze_the_row, freeze_the_col)
            if not flag:
                print("⚠️ Program Stop At `freeze_excel_panes`")
                return False

        if open_file:
            flag = open_file_when_program_end(output_file)
            if not flag:
                print("⚠️ Program Stop At `open_file_when_program_end`")
                return False

        print("✅ 存檔+寫格式 流程完整完成 !!")
    except Exception as e:
        print(f"❌ 流程化處理錯誤: {e}")
        return False


def hide_excel_columns(
        filepath: str, sheet_name: str = "Sheet1",
        hidden_columns_by_name: list[str] = None,
        hidden_columns_by_index: list[int] = None
    ) -> bool:
    """
    開啟現有的 Excel 檔案，並隱藏指定工作表中的欄位 (可依名稱或索引)。

    Args:
        filepath (str): Excel 檔案的路徑。
        sheet_name (str, optional): 要操作的工作表名稱。預設為 "Sheet1"。
        hidden_columns_by_name (list, optional): 需要隱藏的欄位名稱列表。預設為 None。
        hidden_columns_by_index (list, optional): 需要隱藏的欄位索引列表 (從 1 開始)。預設為 None。
    """
    if hidden_columns_by_name is None:
        hidden_columns_by_name = []
    if hidden_columns_by_index is None:
        hidden_columns_by_index = []

    try:
        workbook = load_workbook(filepath)
        if sheet_name not in workbook.sheetnames:
            print(f"❌ 警告：工作表 '{sheet_name}' 不存在於檔案 '{filepath}' 中。")
            return False
        sheet = workbook[sheet_name]
        header = [cell.value for cell in sheet[1]] # 假設第一行為 header
        column_map = {col_name: i + 1 for i, col_name in enumerate(header)}

        # 處理按名稱隱藏
        for col_name_to_hide in hidden_columns_by_name:
            if col_name_to_hide in column_map:
                col_letter = get_column_letter(column_map[col_name_to_hide])
                sheet.column_dimensions[col_letter].hidden = True
            else:
                print(f"❌ 警告：要按名稱隱藏的欄位 '{col_name_to_hide}' 不存在於工作表 '{sheet_name}' 的欄位中。")

        # 處理按索引隱藏
        for col_index_to_hide in hidden_columns_by_index:
            try:
                col_letter = get_column_letter(col_index_to_hide)
                sheet.column_dimensions[col_letter].hidden = True
            except ValueError:
                print(f"❌ 警告：欄位索引 '{col_index_to_hide}' 無效。")

        workbook.save(filepath)
        print(f"⏩ 檔案 '{filepath}', 工作表 '{sheet_name}' 中指定的欄位已隱藏。")
        return True

    except FileNotFoundError:
        print(f"❌ 錯誤：找不到檔案 '{filepath}'，無法隱藏欄位。")
        return False
    except Exception as e:
        print(f"❌ 隱藏檔案 '{filepath}' 中的欄位時發生錯誤：{e}")
        return False

def freeze_excel_panes(
        filepath: str, sheet_name: str = "Sheet1",
        freeze_at_row: int = None, freeze_at_column: int = None
    ) -> bool:
    """
    開啟現有的 Excel 檔案，並凍結指定工作表中的行或列。

    Args:
        filepath (str): Excel 檔案的路徑。
        sheet_name (str, optional): 要操作的工作表名稱。預設為 "Sheet1"。
        freeze_at_row (int, optional): 要凍結的行數 (從 1 開始)，若為 None 則不凍結行。
        freeze_at_column (int, optional): 要凍結的列數 (從 1 開始)，若為 None 則不凍結列。
    """
    try:
        workbook = load_workbook(filepath)
        if sheet_name not in workbook.sheetnames:
            print(f"❌ 警告：工作表 '{sheet_name}' 不存在於檔案 '{filepath}' 中。")
            return False
        sheet = workbook[sheet_name]

        # 設置凍結窗格
        if freeze_at_row and freeze_at_column:
            sheet.freeze_panes = sheet.cell(row=freeze_at_row + 1, column=freeze_at_column + 1)  # 從第 (row+1) 行、第 (column+1) 列凍結
        elif freeze_at_row:
            sheet.freeze_panes = sheet.cell(row=freeze_at_row + 1, column=1)  # 只凍結行
        elif freeze_at_column:
            sheet.freeze_panes = sheet.cell(row=1, column=freeze_at_column + 1)  # 只凍結列

        workbook.save(filepath)
        print(f"⏩ 檔案 '{filepath}', 工作表 '{sheet_name}' 已根據指定凍結行或列。")
        return True

    except FileNotFoundError:
        print(f"❌ 錯誤：找不到檔案 '{filepath}'，無法凍結窗格。")
        return False
    except Exception as e:
        print(f"❌ 凍結檔案 '{filepath}' 中的窗格時發生錯誤：{e}")
        return False

