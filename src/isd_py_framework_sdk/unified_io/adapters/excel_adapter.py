"""
unified_io.adapters.excel_adapter
===================================
Excel read / write adapter backed by pandas + openpyxl.

Requires: ``openpyxl>=3.1``, ``pandas>=2.0``
(``pip install isd-py-framework-sdk[unified_io.excel]``)

Write modes
-----------
``"fresh"`` (default)
    Write via ``pandas.DataFrame.to_excel``.  Fast.  All existing cell
    formatting in the sheet is **lost**.  Use this when setting up a new
    file or when you will apply all formatting afterwards with
    :class:`~isd_py_framework_sdk.excel_painter.ExcelPainter`.

``"preserve"``
    Capture the existing sheet's full format snapshot before writing, then
    restore it after.  Preserves fills, fonts, column widths, freeze panes,
    and merged-cell ranges.  Does **not** preserve CellRichText (the plain
    text version is retained).  Use this when you want to update only the
    cell *values* while keeping a hand-crafted or previously painted layout.

    Write-ordering inside preserve mode::

        1. capture snapshot
        2. clear rows / rewrite data (openpyxl)
        3. restore snapshot (fills, freeze, merges …)
        4. save

``"inplace"``
    Open the existing workbook and update only the cell values in the
    data range — nothing else is touched.  The fastest format-safe option,
    but the DataFrame's shape must match the existing sheet layout
    (same number of columns; rows are written from row 2 onward).
    CellRichText outside the written range is fully preserved.

    Use this when only values change and the sheet structure is stable.

``"styled"`` (or pass ``style=``)
    Bridge to :mod:`isd_py_framework_sdk.excel_painter` — write *and* format
    the table in one step (header band, borders, frozen header, autofilter,
    per-column widths/wrap, optional status colours).  Styling options
    (``widths``, ``wrap_cols``, ``text_cols``, ``status_column``,
    ``status_fills``, ``auto_width`` …) pass straight through.  This is the
    link between the IO layer (data movement) and the painting layer
    (presentation); ``preserve`` mode also relies on excel_painter's
    ``SheetFormatSnapshot``.
"""
from __future__ import annotations

from pathlib import Path
from typing import Literal, Optional
import pandas as pd

from ._interface import IIOAdapter

_MISSING_MSG = (
    "ExcelIOAdapter requires openpyxl and pandas.  "
    "Install with:  pip install isd-py-framework-sdk[unified_io.excel]"
)

WriteMode = Literal["fresh", "preserve", "inplace", "styled"]


class ExcelIOAdapter(IIOAdapter):
    """
    Read / write Excel ``.xlsx`` files.

    Parameters
    ----------
    default_write_mode
        Default write mode (``"fresh"``, ``"preserve"``, or ``"inplace"``).
        Can be overridden per ``write()`` call.
    """

    def __init__(self, default_write_mode: WriteMode = "fresh") -> None:
        self._default_mode = default_write_mode

    # ── Read ─────────────────────────────────────────────────────────────

    def read(
        self,
        source: str | Path,
        *,
        sheet_name: str | int = 0,
        **kwargs,
    ) -> pd.DataFrame:
        """
        Parameters
        ----------
        source
            Path to the ``.xlsx`` file.
        sheet_name
            Sheet name or 0-based index.  Defaults to the first sheet.
        **kwargs
            Forwarded to :func:`pandas.read_excel`.
        """
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise ImportError(_MISSING_MSG) from None
        return pd.read_excel(source, sheet_name=sheet_name, **kwargs)

    # ── Write ─────────────────────────────────────────────────────────────

    def write(
        self,
        df: pd.DataFrame,
        destination: str | Path,
        *,
        sheet_name: str = "Sheet1",
        mode: Optional[WriteMode] = None,
        index: bool = False,
        style=None,
        **kwargs,
    ) -> None:
        """
        Persist *df* to an Excel file.

        Parameters
        ----------
        destination
            Output ``.xlsx`` path.
        sheet_name
            Target worksheet name.
        mode
            Write mode — ``"fresh"``, ``"preserve"``, ``"inplace"``, or
            ``"styled"``.  Overrides the adapter-level default.
        index
            Whether to write the DataFrame index.
        style
            When given (a
            :class:`~isd_py_framework_sdk.excel_painter.TableStyle`, or ``True``
            for the default look), the write is routed through
            :func:`~isd_py_framework_sdk.excel_painter.save_styled_table` to
            produce a professionally formatted table.  Extra styling options
            (``widths``, ``wrap_cols``, ``text_cols``, ``status_column``,
            ``status_fills``, ``auto_width`` …) pass straight through ``**kwargs``.
        **kwargs
            Additional options forwarded to the underlying writer.
        """
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise ImportError(_MISSING_MSG) from None

        effective_mode: WriteMode = mode or self._default_mode
        dest = Path(destination)

        # A supplied style (or mode="styled") bridges into excel_painter.
        if style is not None or effective_mode == "styled":
            self._write_styled(df, dest, sheet_name, index, style, **kwargs)
            return

        if effective_mode == "fresh":
            self._write_fresh(df, dest, sheet_name, index, **kwargs)
        elif effective_mode == "preserve":
            self._write_preserve(df, dest, sheet_name, index)
        elif effective_mode == "inplace":
            self._write_inplace(df, dest, sheet_name)
        else:
            raise ValueError(
                f"Unknown write mode {effective_mode!r}.  "
                "Choose 'fresh', 'preserve', 'inplace', or 'styled'."
            )

    def _write_styled(
        self,
        df: pd.DataFrame,
        dest: Path,
        sheet_name: str,
        index: bool,
        style,
        **kwargs,
    ) -> None:
        """Bridge to excel_painter for a fully styled table write."""
        from ...interop import require_feature

        excel_painter = require_feature("excel_painter")
        TableStyle = excel_painter.TableStyle
        save_styled_table = excel_painter.save_styled_table

        table_style = style if isinstance(style, TableStyle) else None
        save_styled_table(
            df, dest,
            sheet_name=sheet_name,
            style=table_style,
            index=index,
            **kwargs,
        )

    # ── Private write implementations ─────────────────────────────────────

    def _write_fresh(
        self,
        df: pd.DataFrame,
        dest: Path,
        sheet_name: str,
        index: bool,
        **kwargs,
    ) -> None:
        """pandas to_excel — fast, no format preservation."""
        df.to_excel(dest, sheet_name=sheet_name, index=index, **kwargs)

    def _write_preserve(
        self,
        df: pd.DataFrame,
        dest: Path,
        sheet_name: str,
        index: bool,
    ) -> None:
        """
        Snapshot format → clear rows → write data via openpyxl → restore.

        Write-ordering steps performed here:
        1. load workbook
        2. capture SheetFormatSnapshot
        3. delete all data rows (keeps the sheet object)
        4. write header + data rows cell-by-cell
        5. restore snapshot (fills, freeze, merges, column widths)
        6. save
        """
        from ...interop import require_feature

        # Resolve the excel_painter bridge first so a missing openpyxl yields the
        # SDK's standard "pip install ...[excel_painter]" message rather than a
        # bare ImportError from the line below.
        excel_painter = require_feature("excel_painter")
        SheetFormatSnapshot = excel_painter.SheetFormatSnapshot

        from openpyxl import load_workbook

        if not dest.exists():
            # File does not exist yet — fall back to fresh write
            self._write_fresh(df, dest, sheet_name, index)
            return

        wb = load_workbook(dest)
        if sheet_name not in wb.sheetnames:
            # Sheet does not exist — fall back to fresh write on existing book
            with pd.ExcelWriter(dest, engine="openpyxl", mode="a") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
            return

        ws = wb[sheet_name]
        snap = SheetFormatSnapshot.capture(ws)

        # Clear all rows while keeping the sheet structure
        ws.delete_rows(1, ws.max_row)

        # Write header
        all_cols = (
            [df.index.name or "index"] + list(df.columns) if index else list(df.columns)
        )
        for col_idx, col_name in enumerate(all_cols, 1):
            ws.cell(row=1, column=col_idx).value = col_name

        # Write data rows
        iter_df = df.reset_index() if index else df
        for row_offset, record in enumerate(iter_df.itertuples(index=False, name=None), 2):
            for col_idx, val in enumerate(record, 1):
                # Convert numpy / pandas scalar types to Python natives
                ws.cell(row=row_offset, column=col_idx).value = _to_native(val)

        # Restore format (fills, freeze, merges, dimensions)
        snap.restore(ws)

        wb.save(dest)

    def _write_inplace(
        self,
        df: pd.DataFrame,
        dest: Path,
        sheet_name: str,
    ) -> None:
        """
        Open existing workbook and update only cell *values* in the data
        range — all formatting (including CellRichText outside the range)
        is untouched.

        Assumptions
        -----------
        * Row 1 is the header row; data starts at row 2.
        * The DataFrame has the same column count as the existing sheet.
        * Rows beyond ``len(df) + 1`` in the sheet are left as-is.
        """
        from openpyxl import load_workbook

        if not dest.exists():
            self._write_fresh(df, dest, sheet_name, index=False)
            return

        wb = load_workbook(dest)
        if sheet_name not in wb.sheetnames:
            with pd.ExcelWriter(dest, engine="openpyxl", mode="a") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            return

        ws = wb[sheet_name]

        # Update header
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx).value = col_name

        # Update data rows (value only, style untouched)
        for row_offset, record in enumerate(df.itertuples(index=False, name=None), 2):
            for col_idx, val in enumerate(record, 1):
                ws.cell(row=row_offset, column=col_idx).value = _to_native(val)

        wb.save(dest)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _to_native(val):
    """Convert numpy / pandas scalar types to Python-native equivalents."""
    try:
        import numpy as np  # noqa: F401

        if isinstance(val, (float,)) and val != val:  # NaN
            return None
        # numpy integer / float → int / float
        if hasattr(val, "item"):
            return val.item()
    except ImportError:
        pass

    # pandas NA / NaT
    try:
        import pandas as pd

        if pd.isna(val):
            return None
    except (ImportError, TypeError, ValueError):
        pass

    return val
