"""Excel sheet comparison tool.

Config example::

    {
        "target_path": "target.xlsx",
        "bench_path":  "expected.xlsx",
        "sheets": [
            {
                "target_sheet": "Sheet1",
                "bench_sheet": "Sheet1",
                "checks": ["content", "color", "freeze", "hidden"],
                "mask": {
                    "include_rows": [1,2], 
                    "exclude_cells": ["A1"]
                },
                "marker":
                {
                    "known_correct": ["E5"],
                    "known_incorrect": ["N35"]
                }
            }
        ]
    }
"""

from openpyxl import load_workbook
from typing import Dict, Any
from openpyxl.utils import get_column_letter

# color / formatting
GREEN = "\033[92m"
RED = "\033[91m"
YELLOW = "\033[93m"
CYAN = "\033[96m"
RESET = "\033[0m"
PURPLE = "\033[95m"
BOLD = "\033[1m"


def colorize_diff(actual, expected) -> str:
    """將『得到 / 預期』關鍵字上色"""
    return (
        f"{YELLOW}得到{RESET} {actual} | "
        f"{CYAN}預期{RESET} {expected}"
    )


def compare_excel_sheets(config: Dict[str, Any]):
    """
    基於 Config 的 Excel 核對函數
    - 無 skip → 完全維持原行為
    - 有 skip → 啟用 correct-skip / false-skip 統計
    - 有 row_filter / col_filter → 僅檢查指定行/欄（向下相容）
    - 有 mask → 進階 include/exclude 過濾（優先於 row_filter / col_filter）
    """
    target_path = config['target_path']
    bench_path = config['bench_path']

    print(f"\n🚀 {CYAN}開始對照測試...{RESET}")
    print(f"待測檔: {target_path}")
    print(f"標準檔: {bench_path}\n")

    wb_t = load_workbook(target_path, data_only=True)
    wb_b = load_workbook(bench_path, data_only=True)

    # 全域預設要顯示的錯誤行數（可由 config 調整）
    global_max_display = config.get("max_display_errors", 15)

    for sheet_conf in config['sheets']:
        t_name = sheet_conf['target_sheet']
        b_name = sheet_conf['bench_sheet']
        checks = sheet_conf.get('checks', [])

        print(f"📦 檢查工作表: [{t_name}] vs [{b_name}]")

        ws_t = wb_t[t_name]
        ws_b = wb_b[b_name]

        # ---------- marker 設定（可選；舊版 skip 向下相容） ----------
        # marker.known_correct  : 差異是正確的版本更新 → 不列出，計入「已確認正確改動」
        # marker.known_incorrect: 差異已確認為問題   → 不列出，計入「已確認問題」
        marker_conf = sheet_conf.get("marker") or sheet_conf.get("skip")
        if marker_conf:
            # 支援新 key（known_correct / known_incorrect）與舊 key（correct / false）
            correct_skip = set(
                marker_conf.get("known_correct",
                                marker_conf.get("correct", []))
            )
            false_skip = set(
                marker_conf.get("known_incorrect",
                                marker_conf.get("false", []))
            )
        else:
            correct_skip = false_skip = None

        # ---------- row/col filter（可選，向下相容） ----------
        row_filter = sheet_conf.get("row_filter")  # list of int
        col_filter = sheet_conf.get("col_filter")  # list of str (column letters)

        # ---------- mask（優先於 row_filter / col_filter） ----------
        mask_conf = sheet_conf.get("mask")
        if mask_conf:
            _include_rows  = set(mask_conf["include_rows"]) if "include_rows" in mask_conf else None
            _exclude_rows  = set(mask_conf.get("exclude_rows", []))
            _include_cols  = set(mask_conf["include_cols"]) if "include_cols" in mask_conf else None
            _exclude_cols  = set(mask_conf.get("exclude_cols", []))
            _exclude_cells = set(mask_conf.get("exclude_cells", []))
        else:
            _include_rows  = set(row_filter) if row_filter else None
            _exclude_rows  = set()
            _include_cols  = set(col_filter) if col_filter else None
            _exclude_cols  = set()
            _exclude_cells = set()

        # ---------- 統計 ----------
        errors = []
        real_errors = 0
        expected_changes = 0
        false_skips = 0

        # 每個工作表可以覆寫顯示數量
        max_display = sheet_conf.get("max_display_errors", global_max_display)

        # 獲取最大檢查範圍 (以標準檔為準)
        max_row = ws_b.max_row
        max_col = ws_b.max_column

        # ---------- Cell-level checks ----------
        for r in range(1, max_row + 1):
            if _include_rows is not None and r not in _include_rows:
                continue
            if r in _exclude_rows:
                continue
            for c in range(1, max_col + 1):
                col_letter = get_column_letter(c)
                if _include_cols is not None and col_letter not in _include_cols:
                    continue
                if col_letter in _exclude_cols:
                    continue

                cell_t = ws_t.cell(row=r, column=c)
                cell_b = ws_b.cell(row=r, column=c)
                coord = cell_b.coordinate  # A1, B2 ...

                if coord in _exclude_cells:
                    continue

                if correct_skip is not None:
                    is_correct_skip = coord in correct_skip
                    is_false_skip = coord in false_skip
                else:
                    is_correct_skip = is_false_skip = False

                # 1. 檢查內容 (Value)
                if 'content' in checks and cell_t.value != cell_b.value:
                    if is_correct_skip:
                        expected_changes += 1
                        continue
                    if is_false_skip:
                        false_skips += 1
                        continue

                    real_errors += 1
                    errors.append(
                        f"[{coord}] 內容錯誤: "
                        + colorize_diff(cell_t.value, cell_b.value)
                    )

                # 2. 檢查顏色 (Color)
                if 'color' in checks:
                    color_t = str(cell_t.fill.start_color.rgb)[-6:]
                    color_b = str(cell_b.fill.start_color.rgb)[-6:]
                    if color_t != color_b:
                        if is_correct_skip:
                            expected_changes += 1
                            continue
                        if is_false_skip:
                            false_skips += 1
                            continue

                        real_errors += 1
                        errors.append(
                            f"[{coord}] 顏色錯誤: "
                            + colorize_diff(color_t, color_b)
                        )

                # 3. 檢查類型 (Type)
                if 'type' in checks and type(cell_t.value) != type(cell_b.value):
                    if is_correct_skip:
                        expected_changes += 1
                        continue
                    if is_false_skip:
                        false_skips += 1
                        continue

                    real_errors += 1
                    errors.append(
                        f"[{coord}] 類型錯誤: "
                        + colorize_diff(type(cell_t.value), type(cell_b.value))
                    )

        # 4. 檢查凍結視窗 (Freeze Panes)
        if 'freeze' in checks:
            if ws_t.freeze_panes != ws_b.freeze_panes:
                real_errors += 1
                errors.append(
                    "凍結範圍錯誤: "
                    + colorize_diff(ws_t.freeze_panes, ws_b.freeze_panes)
                )

        # 5. 檢查隱藏狀態 (Column Hidden)
        if 'hidden' in checks:
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                if _include_cols is not None and col_letter not in _include_cols:
                    continue
                if col_letter in _exclude_cols:
                    continue
                dim_t = ws_t.column_dimensions.get(col_letter)
                dim_b = ws_b.column_dimensions.get(col_letter)

                status_t = dim_t.hidden if dim_t else False
                status_b = dim_b.hidden if dim_b else False

                if status_t != status_b:
                    real_errors += 1
                    errors.append(
                        f"欄位 [{col_letter}] 隱藏狀態錯誤: "
                        + colorize_diff(status_t, status_b)
                    )

        # ---------- 輸出結果 ----------
        if real_errors == 0:
            msg = f"{GREEN}✓ PASS{RESET}"
            if marker_conf:
                parts = []
                if expected_changes:
                    parts.append(f"{GREEN}{expected_changes} 已確認正確改動{RESET}")
                if false_skips:
                    parts.append(f"{YELLOW}{false_skips} 已確認問題{RESET}")
                if parts:
                    msg += "（" + "，".join(parts) + "）"
            print(f"  {msg}")
        else:
            if marker_conf:
                print(
                    f"  {RED}✗ FAILED{RESET} "
                    f"("
                    f"{GREEN}{expected_changes} 已確認正確改動{RESET} + "
                    f"{YELLOW}{false_skips} 已確認問題{RESET} + "
                    f"{PURPLE}{real_errors} 非預期錯誤{RESET}"
                    f")"
                )
            else:
                print(f"  {RED}✗ FAILED ({real_errors} 個錯誤){RESET}")

            for err in errors[:max_display]:
                print(f"    - {err}")

            if len(errors) > max_display:
                print(f"    ... 以及其餘 {len(errors) - max_display} 個錯誤")

