import pandas as pd
from openpyxl import load_workbook
from typing import Literal, List, Dict, Any

# 顏色定義 (同你的格式)
GREEN = "\033[92m"
RED = "\033[91m"
YELLOW = "\033[93m"
CYAN = "\033[96m"
WHITE = "\033[97m"
PURPLE = "\033[95m"
RESET = "\033[0m"
BOLD = "\033[1m"


def run_excel_structure_test(
        file_path: str,
        expected_sheets: List[str],
        hidden_cols: Dict[str, List[str]],  # { sheet_name: [col_names] }
        color_samples: List[Dict[str, Any]],  # [{ sheet, row, col_name, expected_hex }]
        print_mode: Literal["show_all", "wrong_answer"] = "show_all"
):
    print(f"\n===== Testing Excel Output: {BOLD}{file_path}{RESET} =====")
    wb = load_workbook(file_path)
    all_passed = True

    def log_result(success, message):
        nonlocal all_passed
        if not success: all_passed = False
        show = (print_mode == "show_all") or (not success and print_mode == "wrong_answer")
        if show:
            mark = f"{GREEN}✓{RESET}" if success else f"{RED}✗{RESET}"
            print(f"{mark} {message}")

    # 1. 驗證 Sheet 名稱
    sheet_names = wb.sheetnames
    log_result(sheet_names == expected_sheets, f"Sheet names: {sheet_names} | Expected: {expected_sheets}")

    # 2. 驗證隱藏欄位
    for sheet_name, cols in hidden_cols.items():
        if sheet_name in wb:
            ws = wb[sheet_name]
            header = {cell.value: cell.column for cell in ws[1]}
            for col_name in cols:
                col_idx = header.get(col_name)
                if col_idx:
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    is_hidden = ws.column_dimensions[col_letter].hidden
                    log_result(is_hidden, f"Sheet [{sheet_name}] Column [{col_name}] is hidden")
                else:
                    log_result(False, f"Sheet [{sheet_name}] Column [{col_name}] NOT FOUND")

    # 3. 驗證顏色 (抽樣)
    for sample in color_samples:
        sheet_name = sample['sheet']
        if sheet_name in wb:
            ws = wb[sheet_name]
            header = {cell.value: cell.column for cell in ws[1]}
            col_idx = header.get(sample['col_name'])
            cell = ws.cell(row=sample['row'], column=col_idx)
            # openpyxl 的 rgb 會帶 alpha (如 FFADD8E6)，取後六碼
            actual_rgb = str(cell.fill.start_color.rgb)[-6:]
            expected_rgb = sample['expected_hex'].upper()

            correct = actual_rgb == expected_rgb
            log_result(correct,
                       f"Sheet [{sheet_name}] Row {sample['row']} [{sample['col_name']}] color: {YELLOW}{actual_rgb}{RESET} | Expected: {CYAN}{expected_rgb}{RESET}")

    # 總結
    if all_passed:
        print(f"\n{GREEN}{BOLD}STRUCTURE TEST PASS!{RESET}")
    else:
        print(f"\n{RED}{BOLD}STRUCTURE TEST FAILED!{RESET}")

    return all_passed

