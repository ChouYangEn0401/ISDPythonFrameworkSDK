import pytest
pytest.skip(
    "Deprecated: tests moved to per-format files (test_csv.py, test_json.py, etc.).",
    allow_module_level=True,
)

import os
import csv
import json
import configparser
import xml.etree.ElementTree as ET
import pytest

from tests.file_test.conftest import sample_name, expected_dir, target_dir


def _paths(ext):
    name = sample_name()
    expected = os.path.join(expected_dir(), f"[BU] {name}.{ext}")
    target = os.path.join(target_dir(), f"{name}.{ext}")
    return expected, target


def _exists_or_skip(expected, target):
    if not os.path.exists(expected) or not os.path.exists(target):
        pytest.skip(f"Missing expected/target files: {expected} or {target}")


def test_csv_compare():
    expected, target = _paths("csv")
    _exists_or_skip(expected, target)
    with open(expected, newline='', encoding='utf-8') as fe, open(target, newline='', encoding='utf-8') as ft:
        re = list(csv.reader(fe))
        rt = list(csv.reader(ft))
    assert re == rt


def test_json_compare():
    expected, target = _paths("json")
    _exists_or_skip(expected, target)
    with open(expected, encoding='utf-8') as fe, open(target, encoding='utf-8') as ft:
        je = json.load(fe)
        jt = json.load(ft)
    assert je == jt


def test_jsonl_compare():
    expected, target = _paths("jsonl")
    _exists_or_skip(expected, target)
    with open(expected, encoding='utf-8') as fe, open(target, encoding='utf-8') as ft:
        le = [json.loads(l) for l in fe.read().splitlines() if l.strip()]
        lt = [json.loads(l) for l in ft.read().splitlines() if l.strip()]
    assert le == lt


def test_txt_compare():
    expected, target = _paths("txt")
    _exists_or_skip(expected, target)
    with open(expected, encoding='utf-8') as fe, open(target, encoding='utf-8') as ft:
        te = fe.read()
        tt = ft.read()
    assert te == tt


def test_yaml_compare():
    expected, target = _paths("yaml")
    _exists_or_skip(expected, target)
    yaml = pytest.importorskip("yaml")
    with open(expected, encoding='utf-8') as fe, open(target, encoding='utf-8') as ft:
        ye = yaml.safe_load(fe)
        yt = yaml.safe_load(ft)
    assert ye == yt


def _elements_equal(a, b):
    if a.tag != b.tag:
        return False
    if (a.text or '').strip() != (b.text or '').strip():
        return False
    if (a.tail or '').strip() != (b.tail or '').strip():
        return False
    if a.attrib != b.attrib:
        return False
    if len(a) != len(b):
        return False
    return all(_elements_equal(x, y) for x, y in zip(list(a), list(b)))


def test_xml_compare():
    expected, target = _paths("xml")
    _exists_or_skip(expected, target)
    ee = ET.parse(expected).getroot()
    et = ET.parse(target).getroot()
    assert _elements_equal(ee, et)


def test_ini_compare():
    expected, target = _paths("ini")
    _exists_or_skip(expected, target)
    ce = configparser.ConfigParser()
    ct = configparser.ConfigParser()
    ce.read(expected, encoding='utf-8')
    ct.read(target, encoding='utf-8')
    def as_dict(conf):
        return {s: dict(conf.items(s)) for s in conf.sections()}
    assert as_dict(ce) == as_dict(ct)


def test_toml_compare():
    expected, target = _paths("toml")
    _exists_or_skip(expected, target)
    try:
        import tomllib as toml
    except Exception:
        toml = pytest.importorskip("tomli")
    with open(expected, 'rb') as fe, open(target, 'rb') as ft:
        te = toml.load(fe)
        tt = toml.load(ft)
    assert te == tt


def test_xlsx_compare():
    expected, target = _paths("xlsx")
    _exists_or_skip(expected, target)
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    we = load_workbook(expected, data_only=True)
    wt = load_workbook(target, data_only=True)
    se = we.active
    st = wt.active
    def sheet_to_list(s):
        return [[cell.value for cell in row] for row in s.iter_rows()]
    assert sheet_to_list(se) == sheet_to_list(st)
