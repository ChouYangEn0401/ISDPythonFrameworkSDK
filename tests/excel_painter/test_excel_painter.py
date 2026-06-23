"""
tests/excel_painter/test_excel_painter.py
=========================================

Unit tests for excel_painter.

Test groups
-----------
1.  display_width — CJK-aware widths (+ wcwidth-optional fallback)
2.  highlight — LCS / common-words primitives + CellRichText builders
3.  ExcelPainter — style_table landing, fills, color_scale, number_format
4.  ExcelPainter — auto_width, freeze, hide, title_banner, style_row
5.  SheetFormatSnapshot — capture / rewrite / restore roundtrip
6.  Templates — status_report / summary_statistics_report /
                 multi_sheet_report / diff_highlight_report

Run:
    python -m pytest -v tests/excel_painter/test_excel_painter.py
"""
from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText

from isd_py_framework_sdk.excel_painter import (
    ExcelPainter,
    SheetFormatSnapshot,
    TableStyle,
    STATUS_GREEN,
    STATUS_RED,
    display_width,
    char_lcs,
    word_lcs,
    common_words,
    lcs_rich_text,
    common_words_rich_text,
    auto_status_fills,
    status_report,
    summary_statistics_report,
    multi_sheet_report,
    diff_highlight_report,
)


@pytest.fixture
def df():
    return pd.DataFrame({
        "ISI_ID": ["000123", "000456", "000789"],
        "Name": ["Chen, Shih-Yu", "王小明", "Lee, Min-Hung"],
        "狀態": ["✅完成", "❌失敗", "✅完成"],
        "說明": ["a long english note here", "中文較長的說明文字內容", "another note"],
    })


# ── 1. display_width ──────────────────────────────────────────────────────────

def test_display_width_ascii():
    assert display_width("ABC") == 3
    assert display_width("") == 0
    assert display_width(None) == 0


def test_display_width_cjk():
    # full-width glyphs count as 2
    assert display_width("台大") == 4
    assert display_width("台大ABC") == 7


# ── 2. highlight primitives ───────────────────────────────────────────────────

def test_char_and_word_lcs():
    assert char_lcs("ABCBDAB", "BDCAB") in ("BCAB", "BDAB")  # classic LCS len 4
    assert word_lcs("the quick brown fox", "the brown fox") == ["THE", "BROWN", "FOX"]


def test_common_words():
    assert common_words("a b c", "b c d") == {"b", "c"}
    assert common_words("Hello World", "hello there") == {"hello"}


def test_rich_text_builders():
    rt = lcs_rich_text("National Taiwan University", "Taiwan Univ", mode="word")
    assert isinstance(rt, CellRichText)
    rt2 = common_words_rich_text("Dept Chemistry NTU", "Chemistry National")
    assert isinstance(rt2, CellRichText)


# ── 3. ExcelPainter core ──────────────────────────────────────────────────────

def test_style_table_lands(df, tmp_path):
    p = tmp_path / "t.xlsx"
    (
        ExcelPainter.from_dataframe(df, sheet_name="R")
        .style_table(widths={"Name": 20}, wrap_cols=["說明"], text_cols=["ISI_ID"])
        .save(p)
    )
    ws = load_workbook(p)["R"]
    assert ws["A1"].font.bold is True
    assert ws["A1"].fill.fgColor.rgb.endswith("305496")
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:D4"
    assert ws["A2"].number_format == "@"          # ISI_ID forced text
    assert ws["D2"].alignment.wrap_text is True   # 說明 wraps


def test_fill_by_value(df, tmp_path):
    p = tmp_path / "t.xlsx"
    (
        ExcelPainter.from_dataframe(df)
        .fill_by_value("狀態", {"✅完成": STATUS_GREEN, "❌失敗": STATUS_RED})
        .save(p)
    )
    ws = load_workbook(p).active
    assert ws.cell(row=2, column=3).fill.fgColor.rgb.endswith("C6EFCE")  # green
    assert ws.cell(row=3, column=3).fill.fgColor.rgb.endswith("FFC7CE")  # red


def test_color_scale(tmp_path):
    p = tmp_path / "t.xlsx"
    data = pd.DataFrame({"v": [0, 50, 100]})
    ExcelPainter.from_dataframe(data).color_scale("v", low_color="FFFFFF", high_color="000000").save(p)
    ws = load_workbook(p).active
    low = ws.cell(row=2, column=1).fill.fgColor.rgb     # value 0 → white
    high = ws.cell(row=4, column=1).fill.fgColor.rgb    # value 100 → black
    assert low.endswith("FFFFFF")
    assert high.endswith("000000")


def test_fill_where(df, tmp_path):
    p = tmp_path / "t.xlsx"
    (
        ExcelPainter.from_dataframe(df)
        .fill_where("狀態", lambda v: v == "❌失敗", "FFC7CE")
        .save(p)
    )
    ws = load_workbook(p).active
    assert ws.cell(row=3, column=3).fill.fgColor.rgb.endswith("FFC7CE")  # ❌失敗 row
    assert ws.cell(row=2, column=3).fill.patternType in (None, "none")   # ✅完成 untouched


def test_lcs_letter_mode_highlight(tmp_path):
    p = tmp_path / "t.xlsx"
    data = pd.DataFrame({"a": ["ABCDEF"], "b": ["ACE"]})
    (
        ExcelPainter.from_dataframe(data)
        .highlight_lcs("a", "b", mode="letter")
        .save(p)
    )
    ws = load_workbook(p, rich_text=True).active
    assert isinstance(ws["A2"].value, CellRichText)


def test_banded_table(df, tmp_path):
    p = tmp_path / "t.xlsx"
    ExcelPainter.from_dataframe(df).style_table(TableStyle(banded=True)).save(p)
    ws = load_workbook(p).active
    # zebra stripes every other data row; row 3 (2nd data row) carries the band
    assert ws.cell(row=3, column=1).fill.fgColor.rgb.endswith("F2F2F2")
    assert ws.cell(row=2, column=1).fill.patternType in (None, "none")


def test_column_ref_by_name_and_index(df, tmp_path):
    painter = ExcelPainter.from_dataframe(df)
    assert painter._col("狀態") == 3
    assert painter._col(3) == 3
    with pytest.raises(KeyError):
        painter._col("does-not-exist")


# ── 4. structure helpers ──────────────────────────────────────────────────────

def test_title_banner_and_style_row(df, tmp_path):
    p = tmp_path / "t.xlsx"
    painter = ExcelPainter.new("S")
    painter.title_banner("My Report", n_cols=len(df.columns))
    painter.write_dataframe(df, start_row=2)
    painter.style_table(header_row=2)
    painter.style_row(2, fill_color="DDEBF7", bold=True)
    painter.save(p)
    ws = load_workbook(p)["S"]
    assert ws["A1"].value == "My Report"
    assert "A1:D1" in [str(r) for r in ws.merged_cells.ranges]


def test_hide_and_auto_width(df, tmp_path):
    p = tmp_path / "t.xlsx"
    (ExcelPainter.from_dataframe(df).auto_width().hide_columns(by_name=["ISI_ID"]).save(p))
    ws = load_workbook(p).active
    assert ws.column_dimensions["A"].hidden is True
    assert ws.column_dimensions["B"].width and ws.column_dimensions["B"].width > 0


# ── 5. snapshot roundtrip ─────────────────────────────────────────────────────

def test_snapshot_restore(df, tmp_path):
    p = tmp_path / "t.xlsx"
    painter = ExcelPainter.from_dataframe(df).style_table().fill_cell(2, 1, "FFFF00")
    snap = painter.snapshot()
    before = painter.ws.cell(row=2, column=1).fill.fgColor.rgb
    # destroy + rewrite values
    painter.ws.delete_rows(1, painter.ws.max_row)
    painter.ws["A1"] = "x"
    painter.restore(snap)
    after = painter.ws.cell(row=2, column=1).fill.fgColor.rgb
    assert before == after
    assert painter.ws["A1"].fill.fgColor.rgb.endswith("305496")  # header band restored


# ── 6. templates ──────────────────────────────────────────────────────────────

def test_auto_status_fills():
    m = auto_status_fills(["✅ 已回填", "⚠️ 需人工", "❌ 失敗", "其他"])
    assert m["✅ 已回填"] == STATUS_GREEN
    assert m["⚠️ 需人工"][0] == "FFEB9C"
    assert m["❌ 失敗"] == STATUS_RED


def test_status_report(df, tmp_path):
    p = status_report(df, tmp_path / "sr.xlsx", status_column="狀態",
                      title="T", text_cols=["ISI_ID"], wrap_cols=["說明"])
    wb = load_workbook(p)
    assert "結果" in wb.sheetnames and "統計摘要" in wb.sheetnames
    ws = wb["結果"]
    assert ws["A1"].value == "T"                    # title banner
    # status colour landed on the 狀態 column (col 3, data starts row 3 under title)
    assert ws.cell(row=3, column=3).fill.fgColor.rgb.endswith("C6EFCE")


def test_summary_statistics_report(tmp_path):
    data = pd.DataFrame({"系所": ["A", "B", "C"], "2023": [10, 20, 30], "2024": [40, 50, 60]})
    p = summary_statistics_report(data, tmp_path / "stats.xlsx",
                                  numeric_cols=["2023", "2024"], label_col="系所", title="S")
    ws = load_workbook(p).active
    total_row = 2 + len(data) + 1  # header_row(2) + 3 data + 1
    assert ws.cell(row=total_row, column=1).value == "總計"
    assert ws.cell(row=total_row, column=2).value == 60   # sum of 2023
    assert ws.cell(row=total_row, column=3).value == 150  # sum of 2024


def test_multi_sheet_report(tmp_path):
    a = pd.DataFrame({"x": [1, 2]})
    b = pd.DataFrame({"y": [3, 4, 5]})
    p = multi_sheet_report({"A": a, "B": b}, tmp_path / "ms.xlsx")
    wb = load_workbook(p)
    assert wb.sheetnames[0] == "總覽"              # overview moved to front
    assert set(wb.sheetnames) == {"總覽", "A", "B"}
    ov = wb["總覽"]
    # overview lists row counts (title row1, header row2, data row3+)
    assert ov.cell(row=3, column=2).value in (2, 3)


def test_diff_highlight_report(tmp_path):
    data = pd.DataFrame({
        "a": ["Dept Chem National Taiwan Univ"],
        "b": ["Chemistry Taiwan University"],
    })
    p = diff_highlight_report(data, tmp_path / "diff.xlsx",
                              column_pairs=[("a", "b")], method="common_words")
    ws = load_workbook(p, rich_text=True).active
    assert isinstance(ws["A2"].value, CellRichText)


if __name__ == "__main__":
    import sys
    sys.exit(pytest.main([__file__, "-v"]))
